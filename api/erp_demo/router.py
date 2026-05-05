# -*- coding: utf-8 -*-
from __future__ import annotations

import io
from typing import Any
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from api.erp_demo import models
from api.erp_demo.deps import DbSession
from api.erp_demo.odoo_auth import try_login
from api.erp_demo.schemas import (
    ContractCreate,
    ContractOut,
    CustomerCreate,
    CustomerOut,
    DashboardOut,
    DocumentCreate,
    DocumentOut,
    ErpChatRequest,
    LoginRequest,
    LoginResponse,
    NotificationCreate,
    NotificationOut,
)
from api.erp_demo.security import TokenPayload, build_payload_after_login, create_access_token, get_current_user

router = APIRouter(prefix="/api/erp", tags=["erp-demo"])


def _visible_uid(user: TokenPayload):
    if user.role == "admin":
        return None
    return user.uid


@router.get("/health")
def erp_health() -> dict[str, str]:
    return {"status": "ok", "module": "erp-demo"}


@router.post("/auth/login", response_model=LoginResponse)
def erp_login(body: LoginRequest) -> LoginResponse:
    try:
        uid, display, company, db = try_login(body.login.strip(), body.password)
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Không kết nối được Odoo: {e!s}") from e
    payload = build_payload_after_login(body.login.strip(), uid, db, display, company)
    token = create_access_token(payload)
    return LoginResponse(
        access_token=token,
        role=payload.role,
        name=payload.name,
        company=payload.company,
    )


@router.get("/me")
def erp_me(user: TokenPayload = Depends(get_current_user)) -> dict[str, Any]:
    return {
        "login": user.sub,
        "name": user.name,
        "company": user.company,
        "role": user.role,
        "odoo_uid": user.uid,
        "odoo_db": user.db,
    }


def _dashboard(db: Session, user: TokenPayload) -> DashboardOut:
    uid = _visible_uid(user)

    def count_q(model):
        q = select(func.count()).select_from(model)
        if uid is not None:
            q = q.where(model.owner_odoo_uid == uid)
        return int(db.execute(q).scalar_one())

    cq = select(models.Customer).order_by(models.Customer.created_at.desc()).limit(5)
    if uid is not None:
        cq = cq.where(models.Customer.owner_odoo_uid == uid)
    recent_cust = list(db.execute(cq).scalars())

    ctq = select(models.Contract).order_by(models.Contract.created_at.desc()).limit(5)
    if uid is not None:
        ctq = ctq.where(models.Contract.owner_odoo_uid == uid)
    recent_contracts = list(db.execute(ctq).scalars())

    nq = select(func.count()).select_from(models.Notification).where(models.Notification.is_read.is_(False))
    if uid is not None:
        nq = nq.where(models.Notification.owner_odoo_uid == uid)
    unread = int(db.execute(nq).scalar_one())

    return DashboardOut(
        counts={
            "customers": count_q(models.Customer),
            "contracts": count_q(models.Contract),
            "documents": count_q(models.Document),
            "notifications": count_q(models.Notification),
        },
        recent_customers=[CustomerOut.model_validate(x) for x in recent_cust],
        recent_contracts=[ContractOut.model_validate(x) for x in recent_contracts],
        unread_notifications=unread,
    )


@router.get("/dashboard", response_model=DashboardOut)
def erp_dashboard(db: DbSession, user: TokenPayload = Depends(get_current_user)) -> DashboardOut:
    return _dashboard(db, user)


# --- Khách hàng ---
@router.get("/customers", response_model=list[CustomerOut])
def list_customers(db: DbSession, user: TokenPayload = Depends(get_current_user)) -> list[CustomerOut]:
    q = select(models.Customer).order_by(models.Customer.created_at.desc())
    uid = _visible_uid(user)
    if uid is not None:
        q = q.where(models.Customer.owner_odoo_uid == uid)
    rows = db.execute(q).scalars().all()
    return [CustomerOut.model_validate(r) for r in rows]


@router.post("/customers", response_model=CustomerOut)
def create_customer(
    body: CustomerCreate,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> CustomerOut:
    row = models.Customer(
        name=body.name.strip(),
        tax_id=(body.tax_id or "").strip() or None,
        email=(body.email or "").strip() or None,
        phone=(body.phone or "").strip() or None,
        address=(body.address or "").strip() or None,
        status=body.status.strip() or "active",
        owner_odoo_uid=user.uid,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return CustomerOut.model_validate(row)


@router.delete("/customers/{cid}")
def delete_customer(
    cid: UUID,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> dict[str, bool]:
    row = db.get(models.Customer, cid)
    if not row:
        raise HTTPException(404, "Không tìm thấy.")
    if user.role != "admin" and row.owner_odoo_uid != user.uid:
        raise HTTPException(403, "Không có quyền xóa.")
    db.delete(row)
    db.commit()
    return {"ok": True}


# --- Hợp đồng ---
@router.get("/contracts", response_model=list[ContractOut])
def list_contracts(db: DbSession, user: TokenPayload = Depends(get_current_user)) -> list[ContractOut]:
    q = select(models.Contract).order_by(models.Contract.created_at.desc())
    uid = _visible_uid(user)
    if uid is not None:
        q = q.where(models.Contract.owner_odoo_uid == uid)
    rows = db.execute(q).scalars().all()
    return [ContractOut.model_validate(r) for r in rows]


@router.post("/contracts", response_model=ContractOut)
def create_contract(
    body: ContractCreate,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> ContractOut:
    cust = db.get(models.Customer, body.customer_id)
    if not cust:
        raise HTTPException(422, "Khách hàng không tồn tại.")
    if user.role != "admin" and cust.owner_odoo_uid != user.uid:
        raise HTTPException(403, "Không gắn hợp đồng với khách của người khác.")
    row = models.Contract(
        customer_id=body.customer_id,
        title=body.title.strip(),
        contract_type=body.contract_type.strip(),
        status=body.status.strip(),
        start_date=body.start_date,
        end_date=body.end_date,
        amount=body.amount,
        notes=(body.notes or "").strip() or None,
        owner_odoo_uid=user.uid,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return ContractOut.model_validate(row)


@router.delete("/contracts/{eid}")
def delete_contract(
    eid: UUID,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> dict[str, bool]:
    row = db.get(models.Contract, eid)
    if not row:
        raise HTTPException(404, "Không tìm thấy.")
    if user.role != "admin" and row.owner_odoo_uid != user.uid:
        raise HTTPException(403, "Không có quyền xóa.")
    db.delete(row)
    db.commit()
    return {"ok": True}


# --- Tài liệu ---
@router.get("/documents", response_model=list[DocumentOut])
def list_documents(db: DbSession, user: TokenPayload = Depends(get_current_user)) -> list[DocumentOut]:
    q = select(models.Document).order_by(models.Document.created_at.desc())
    uid = _visible_uid(user)
    if uid is not None:
        q = q.where(models.Document.owner_odoo_uid == uid)
    rows = db.execute(q).scalars().all()
    return [DocumentOut.model_validate(r) for r in rows]


@router.post("/documents", response_model=DocumentOut)
def create_document(
    body: DocumentCreate,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> DocumentOut:
    if body.customer_id:
        c = db.get(models.Customer, body.customer_id)
        if not c:
            raise HTTPException(422, "Khách hàng không tồn tại.")
        if user.role != "admin" and c.owner_odoo_uid != user.uid:
            raise HTTPException(403, "Không liên kết tài liệu tới khách của người khác.")
    if body.contract_id:
        e = db.get(models.Contract, body.contract_id)
        if not e:
            raise HTTPException(422, "Hợp đồng không tồn tại.")
        if user.role != "admin" and e.owner_odoo_uid != user.uid:
            raise HTTPException(403, "Không liên kết tài liệu tới hợp đồng của người khác.")
    row = models.Document(
        title=body.title.strip(),
        doc_type=body.doc_type.strip(),
        external_url=(body.external_url or "").strip() or None,
        notes=(body.notes or "").strip() or None,
        customer_id=body.customer_id,
        contract_id=body.contract_id,
        owner_odoo_uid=user.uid,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return DocumentOut.model_validate(row)


@router.delete("/documents/{did}")
def delete_document(
    did: UUID,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> dict[str, bool]:
    row = db.get(models.Document, did)
    if not row:
        raise HTTPException(404, "Không tìm thấy.")
    if user.role != "admin" and row.owner_odoo_uid != user.uid:
        raise HTTPException(403, "Không có quyền xóa.")
    db.delete(row)
    db.commit()
    return {"ok": True}


# --- Thông báo ---
@router.get("/notifications", response_model=list[NotificationOut])
def list_notifications(db: DbSession, user: TokenPayload = Depends(get_current_user)) -> list[NotificationOut]:
    q = select(models.Notification).order_by(models.Notification.created_at.desc())
    uid = _visible_uid(user)
    if uid is not None:
        q = q.where(models.Notification.owner_odoo_uid == uid)
    rows = db.execute(q).scalars().all()
    return [NotificationOut.model_validate(r) for r in rows]


@router.post("/notifications", response_model=NotificationOut)
def create_notification(
    body: NotificationCreate,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> NotificationOut:
    row = models.Notification(
        title=body.title.strip(),
        body=(body.body or "").strip() or None,
        due_at=body.due_at,
        priority=body.priority.strip() or "normal",
        owner_odoo_uid=user.uid,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return NotificationOut.model_validate(row)


@router.patch("/notifications/{nid}/read")
def mark_read(
    nid: UUID,
    db: DbSession,
    user: TokenPayload = Depends(get_current_user),
) -> dict[str, bool]:
    row = db.get(models.Notification, nid)
    if not row:
        raise HTTPException(404, "Không tìm thấy.")
    if user.role != "admin" and row.owner_odoo_uid != user.uid:
        raise HTTPException(403, "Không có quyền.")
    row.is_read = True
    db.commit()
    return {"ok": True}


# --- Chatbot có ngữ cảnh ---
@router.post("/chat")
def erp_chat(
    body: ErpChatRequest,
    user: TokenPayload = Depends(get_current_user),
):
    from api.main import ChatRequest, chat as rag_chat

    ctx_lines = [
        "[Ngữ cảnh doanh nghiệp — chỉ để trợ lý hiểu ai đang hỏi, không thay thế văn bản pháp luật]",
        f"Công ty / tổ chức (theo Odoo): {user.company or '—'}",
        f"Người dùng: {user.name} (đăng nhập: {user.sub})",
        f"Database Odoo: {user.db}",
    ]
    prefix = "\n".join(ctx_lines)
    q = f"{prefix}\n\nCâu hỏi pháp lý:\n{body.question.strip()}"
    req = ChatRequest(question=q, top_k=body.top_k, business_groups=body.business_groups)
    return rag_chat(req)


# --- Xuất Excel / PDF ---
@router.get("/export/customers.xlsx")
def export_customers_xlsx(db: DbSession, user: TokenPayload = Depends(get_current_user)):
    q = select(models.Customer).order_by(models.Customer.name)
    uid = _visible_uid(user)
    if uid is not None:
        q = q.where(models.Customer.owner_odoo_uid == uid)
    rows = db.execute(q).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Khach hang"
    ws.append(["Ten", "MST", "Email", "Dien thoai", "Trang thai"])
    for r in rows:
        ws.append([r.name, r.tax_id or "", r.email or "", r.phone or "", r.status])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="khach-hang.xlsx"'},
    )


@router.get("/export/dashboard.pdf")
def export_dashboard_pdf(db: DbSession, user: TokenPayload = Depends(get_current_user)):
    dash = _dashboard(db, user)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_font("Helvetica", size=12)
    pdf.cell(0, 10, txt="Luat Mai Trang - Dashboard demo (PDF)", ln=True)
    pdf.ln(4)
    for k, v in dash.counts.items():
        pdf.cell(0, 8, txt=f"{k}: {v}", ln=True)
    pdf.ln(4)
    pdf.cell(0, 8, txt=f"Unread notifications: {dash.unread_notifications}", ln=True)
    raw = pdf.output(dest="S")
    data = raw if isinstance(raw, (bytes, bytearray)) else raw.encode("latin-1")
    return Response(
        content=bytes(data),
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="dashboard-demo.pdf"'},
    )
